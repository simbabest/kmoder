# -*- coding: utf-8 -*-
"""HuangXin"""
from datetime import date
import glob
import json
import os
import sys
from PyQt5.QtCore import QThread, pyqtSignal, pyqtSlot, Qt
from PyQt5.QtWidgets import (QApplication, QDialog, QFileDialog,
                             QMessageBox, QTableWidgetItem)
from PyQt5.QtGui import QIcon, QColor, QBrush
from openpyxl import Workbook
import xlrd

from ui_wh import Ui_WH

# from openpyxl.styles import Color, Font, Alignment
# from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW
OUTPUT = '_output_'
CALE = 'cale.json'
USECOL = 'A,B,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,A,A,AB,AC,AD,AE,AF,AG,AH,AJ,AK'


def cov_float_to_hhmm(dig):
    """cov float to hhmm"""
    hi = int(dig * 24.0)
    hh = '0' + str(hi) if hi < 10 else str(hi)
    m = int((dig * 24.0 - int(dig * 24.0)) * 60.00)
    mm = '0' + str(m) if m < 10 else str(m)
    return hh + ':' + mm


def calc_work_hour(ss, ee):
    if ee > ss:
        if ss >= 0.5 or (ee > 0.45 and ee <= 0.5):
            rtn = (ee - ss) * 24.00
        else:
            rtn = (ee - ss) * 24.00 - 2.0
    else:
        rtn = (1.0 + ee - ss) * 24.00 - 2.0
    return rtn


class WHS(QDialog, Ui_WH):
    """ WHS """
    def __init__(self, parent=None):
        QDialog.__init__(self, parent)
        self.setupUi(self)
        self.worker = XlsReader()
        self.connect_signal()
        self.dirs = '##'
        self.cale = None
        self.yyyy = date.today().strftime('%Y')
        self.m = int(date.today().strftime('%m'))
        self.d = int(date.today().strftime('%d'))
        self.init_ui_params()
        self.get_cale()

    def get_cale(self):
        """get_cale"""
        try:
            fp = open(CALE, 'r')
        except Exception as _:
            QMessageBox.information(self, "读取失败", "无法读取{cale}配置文件".format(cale=CALE))
        else:
            self.cale = json.loads(fp.read())

    def init_ui_params(self):
        self.icon = QIcon("d13.png")
        self.setWindowIcon(self.icon)
        cnt = 0
        # print('self.d=',self.d)
        hfmlist = ['_1', ''] if (self.d > 8 and self.d < 21) else ['', '_1']
        for m in sorted(range(self.m), reverse=True):
            mon = '0' + str(m + 1) if m + 1 < 10 else str(m + 1)
            for hfm in hfmlist:
                self.cbx_months.addItem(self.yyyy + mon + hfm)
                # self.cbx_months.setItemText(cnt,self.yyyy + mon + hfm)
                cnt += 1
        self.btn_run.setEnabled(False)

    def connect_signal(self):
        self.worker.sig_fi.connect(self.UpText)
        self.worker.tb_data.connect(self.showTable)
        self.worker.err.connect(self.errMsg)
        self.worker.finished.connect(self.Enable_btn)

    @pyqtSlot()
    def on_btn_brsw_clicked(self):
        # self.outputWidget.setText(str(value + self.inputSpinBox2.value()))
        # print('browser path')
        mainPath = self.lineEdit_browser.text()
        # print(mainPath)
        dbfn = QFileDialog.getExistingDirectory(None, u"请选择数据所在目录",
                                                mainPath, QFileDialog.ShowDirsOnly |
                                                QFileDialog.HideNameFilterDetails | QFileDialog.ReadOnly)
        if dbfn:
            self.dirs = dbfn
            self.lineEdit_browser.setText(dbfn)
            group = dbfn.split('/')[-1]
            for m in range(12):
                mon = '0' + str(m + 1) if m + 1 < 10 else str(m + 1)
                for hfm in ['', '_1']:
                    monfolder = self.yyyy + mon + hfm
                    # print(monfolder)
                    try:
                        os.makedirs(dbfn + '/' + monfolder)
                    except Exception as _:
                        pass
                    else:
                        print('make dir success:', monfolder)
            try:
                os.makedirs(dbfn + '/' + OUTPUT)
            except Exception as _:
                pass
            else:
                print('make out dir success:')
            self.label_group.setText(group)
            self.btn_run.setEnabled(True)
        else:
            print('Nothing to do')

    @pyqtSlot()
    def on_btn_run_clicked(self):
        # self.outputWidget.setText(str(value + self.inputSpinBox2.value()))
        self.btn_run.setEnabled(False)
        # print('run')
        # print('dir=', self.dirs)
        curmon = self.cbx_months.currentText()
        self.worker.render(dirs=self.dirs, mon=curmon, cale=self.cale)

    def showTable(self, tb_data):
        tbdata = json.loads(tb_data)
        print(tbdata)
        # cols = len(tbdata['header'])
        cols = len(tbdata['tbody'][0])    
        rows = len(tbdata['tbody'])
        self.tableWidget.setColumnCount(cols)
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setHorizontalHeaderLabels(tbdata['header'])
        for i in range(self.tableWidget.rowCount()):
            for j in range(self.tableWidget.columnCount()):
                cnt = tbdata['tbody'][i][j]
                if i % 3 == 2:
                    if type(cnt) == float:
                        newItem = QTableWidgetItem(str(round(cnt, 2)))
                    else:
                        newItem = QTableWidgetItem(str(cnt))
                else:
                    if type(cnt) == float:
                        newItem = QTableWidgetItem(cov_float_to_hhmm(cnt))
                    else:
                        newItem = QTableWidgetItem(str(cnt))
                if j == self.tableWidget.columnCount() - 1 and i % 3 == 2:
                    brush = QBrush(QColor(200, 139, 33))
                    brush.setStyle(Qt.SolidPattern)
                    newItem.setBackground(brush)
                    brush = QBrush(QColor(24, 37, 216))
                    brush.setStyle(Qt.SolidPattern)
                    newItem.setForeground(brush)
                self.tableWidget.setItem(i, j, newItem)
        self.tableWidget.setColumnWidth(0, 50)
        self.tableWidget.setColumnWidth(1, 40)
        self.tableWidget.setColumnWidth(self.tableWidget.columnCount() - 1, 50)
        for i in range(cols - 3):
            self.tableWidget.setColumnWidth(i + 2, 40)
        for spr in range(int(rows / 3)):
            self.tableWidget.setSpan(3 * spr, 0, 3, 1)

    def UpText(self, zz):
        self.label_savedata.setText(zz)

    def Enable_btn(self):
        self.btn_run.setEnabled(True)

    def errMsg(self, msg):
        self.tableWidget.clear()
        QMessageBox.information(self, "错误消息", msg)


class XlsReader(QThread):
    sig_do = pyqtSignal(int, name='sig_do')
    sig_fi = pyqtSignal(str, name='sig_fi')
    sig_ok = pyqtSignal(int, str, name='sig_ok')
    tb_data = pyqtSignal(str, name='tb_data')
    err = pyqtSignal(str, name='err')
    finished = pyqtSignal(name='finished')

    def __init__(self, parent=None):
        QThread.__init__(self, parent)
        self.exiting = False
        self.idx = 0
        self.dirs = ''
        self.mon = ''
        self.cale = {}

    def __del__(self):
        self.exiting = True
        # self.wait()

    def render(self, dirs, mon, cale):
        self.dirs = dirs
        self.mon = mon
        self.cale = cale
        self.start()

    def read_xlsx(self, mon, xls):
        fnname = xls.split('.')
        name = fnname[0].split('_')[-1]
        wb = xlrd.open_workbook(filename=xls)
        sht = wb.sheet_by_index(0)
        dtkeys = []
        for crange in sorted(sht.merged_cells):
            rs, re, cs, ce = crange
            if re - rs == 1 and ce - cs == 2:
                dt = sht.cell_value(rs, cs)
                if dt:
                    dtkeys.append({'dt': dt, 'cs': cs, 'rs': rs})
        print('len(dtkeys)=', len(dtkeys))
        dtrs = [dts['rs'] for dts in dtkeys]
        dtrs = sorted(list(set(dtrs)))
        rloc = []
        for rid in range(sht.nrows):
            rowv = sht.row_values(rid)
            if rowv[0] == name:
                rloc.append(rid)
        # print(dtkeys)
        result = []
        rsplit = dtrs[1] - dtrs[0]
        prer = rloc[0] - rsplit if rloc[0] > rsplit else 0
        for i in range(len(rloc)):
            rid = rloc[i]
            for dt in dtkeys:
                dtlabel = dt['dt']
                cs = dt['cs']
                rs = dt['rs']
                if rs > prer and rs < rid:
                    ss = sht.cell_value(rid, cs)
                    ee = sht.cell_value(rid, cs + 1)
                    se = 0.00
                    if ee:
                        se = calc_work_hour(ss, ee)

                    d = {'dt': dtlabel, 'ss': ss, 'ee': ee, 'se': se}
                    result.append(d)
            prer = rid
        blank = len(dtkeys) - len(result)
        print('len(result)', len(result))
        if blank > 0:
            print('新入职')
            temp = {'dt': '', 'ss': '', 'ee': '', 'se': 0.00}
            for i in range(blank):
                result.insert(0, temp.copy())
        return {'name': name, 'data': result}

    def write_xlsx(self, mon, data):
        # open workbook and create sheet
        tbdata = {'header': None, 'tbody': []}
        wb = Workbook()
        ws = wb.active
        ws.title = "WorkHour"
        hm = mon[4:6]
        header = ['姓名', str(int(hm)) + '月']
        for dt in self.cale.keys():
            dts = dt.split('-')
            if dts[1] == hm:
                header.append(str(int(dts[2])) + '号')
        header.append('累计')
        tbdata['header'] = header
        ws.append(header)
        # body with for and merge
        ix = 2
        for ir, item in enumerate(data):
            name = item['name']
            sb = [name, '上班']
            xb = [name, '下班']
            sc = [name, '时长']
            totalsc = 0.00
            for day in item['data']:
                sb.append(day['ss'])
                xb.append(day['ee'])
                sc.append(day['se'])
                totalsc += day['se']
            sb.append('')
            xb.append('')
            sc.append(totalsc)
            for line in [sb, xb, sc]:
                ws.append(line)
                tbdata['tbody'].append(line)
                if ix % 3 == 1:
                    # print('ix=', ix, 'set 0.00')
                    for cell in ws["{r}:{r}".format(r=ix)]:
                        cell.number_format = '0.00'
                    # ws.row_dimensions[ix].number_format = '0.00'
                else:
                    # print('ix=', ix, 'set h:mm')
                    for cell in ws["{r}:{r}".format(r=ix)]:
                        cell.number_format = 'h:mm'
                ix += 1
            ws.merge_cells('A{}:A{}'.format((1 + ir) * 3 - 1, (1 + ir) * 3 + 1))
        # ws['C2:AH2'].number_format ='h:mm'

        for ic in USECOL.split(','):
            ws.column_dimensions[ic].width = 6
        fn = self.dirs + '/_output_/{mon}.xlsx'.format(mon=mon)
        try:
            wb.save(fn)
        except Exception as _:
            self.err.emit('请关闭输出的Excel文件,然后重试')
        else:
            self.sig_fi.emit(fn)
            self.tb_data.emit(json.dumps(tbdata))   

    def run(self):
        datadir = self.dirs + '/' + self.mon
        xlsfn = glob.glob(datadir + '/*.xlsx')
        data = []
        for xls in xlsfn:
            xx = xls.split('\\')
            if not xx[-1].startswith('~'):
                wk = self.read_xlsx(self.mon, xls)
                data.append(wk)
        self.write_xlsx(mon=self.mon, data=data)
        self.finished.emit()


def raw_main():
    app = QApplication(sys.argv)
    win = QDialog()
    ui = Ui_WH()
    ui.setupUi(win)
    win.show()
    sys.exit(app.exec_())


def main():
    app = QApplication(sys.argv)
    win = WHS()
    win.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
